from cv2 import cv2
import numpy as np
import configuracion
import detector_caras
import pandas as pd
from openpyxl import Workbook, load_workbook 
ncara=configuracion.nombre_caras()
colores_puntos=configuracion.colores_finales()
#Mando a llamar del archivo configuracion los matices a detectar
#R2_rojo,R_azul,R_naranja,R_verde, R_amarillo, R_blanco=configuracion.matices()
Azul, Verde, Rojo, Naranja,Amarillo, Blanco=configuracion.colores_basic()

def SaveHSV():
    wb=load_workbook('Base_Colores.xlsx')
    ws=wb['Recepcion']
    indice=1136
    for c in range(0,6):
        cara=ncara[c]
        for npunto in range(1,10):
            punto=cv2.imread('Fotos_caras\Puntos\cara_'+ncara[c]+'\C_'+ncara[c]+'_P_'+str(npunto)+'.jpg') 
            #punto=cv2.imread('Fotos_caras\Puntos\cara_'+ncara[c]+'\C_'+ncara[c]+'_P_'+str(npunto)+'.jpg') 
            #frameHSV = cv2.cvtColor(punto, cv2.COLOR_BGR2HSV)
            colors, count = np.unique(punto.reshape(-1, punto.shape[-1]), axis=0, return_counts=True)
            #Primer_color=colors[np.argsort(-count)][1988:1990]
            Primer_color=colors[np.argsort(-count)][:5]
            for i in range(5):
                B=Primer_color[i][0]
                G=Primer_color[i][1]
                R=Primer_color[i][2]
                ws['A'+str(indice)]=B
                ws['B'+str(indice)]=G
                ws['C'+str(indice)]=R
                ws['F'+str(indice)]=colores_puntos[c]
                indice=indice+1
            #print(Primer_color)
            #ws['G'+str(indice)]=colores_puntos[c][npunto]
            #ws['G'+str(indice)]=colores_puntos[c]
            wb.save('Base_Colores.xlsx')

#SaveHSV()

