from cv2 import cv2
import numpy as np
import configuracion
import caras
import pandas as pd
from openpyxl import Workbook, load_workbook 


wb=load_workbook('Base_Colores.xlsx')
ws=wb['RGB']
indice=8
punto=cv2.imread("azul.jpg") 
frameHSV = cv2.cvtColor(punto, cv2.COLOR_BGR2HSV)
colors, count = np.unique(punto.reshape(-1, punto.shape[-1]), axis=0, return_counts=True)
Primer_color=colors[np.argsort(-count)][:5]
for i in range(5):
    B=Primer_color[i][0]
    G=Primer_color[i][1]
    R=Primer_color[i][2]
    ws['A'+str(indice)]=B
    ws['B'+str(indice)]=G
    ws['C'+str(indice)]=R
#ws['G'+str(indice)]=colores_puntos[c][npunto]
#ws['G'+str(indice)]=colores_puntos[c]
wb.save('Base_Colores.xlsx')